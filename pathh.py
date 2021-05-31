import openpyxl
from openpyxl.utils import get_column_letter
import os

path = 'C:\\Users\\Arkadiusz\\Desktop\\FV\\Maj'


cur = os.getcwd()                # Check current working directory.
print('Current working directory %s' % cur)  # Print current working directory in console

os.chdir(path)                  # Change the directory
cur2 = os.getcwd()               # Check current working directory
print(cur2)                     # Print current working directory in console

print('-----')                  # Print ------
wb = openpyxl.load_workbook('domeny-tf.xlsx') # Load excell form path
# sheet = wb['domeny-tf']         # Retrieving a sheet from a workbook

sheet = wb.active



# row = sheet.max_row             # Print maximum number of occupied row in console
# print(row)                                  # print row
# print(sheet.max_column)                     # print maximum number of occupied column
# print(get_column_letter(sheet.max_column))  # print maximum column letter of occupied

# for i in range(1,row +1):
#     print(i, sheet.cell(row=i, column=1).value)

# print(tuple(sheet['A1':'C3']))

# for rowOfCellObjects in sheet['A1':'C3']:   # Iteraction over the rows
#     for cellObj in rowOfCellObjects:        # Iteraction over the column
#         print(cellObj.coordinate, cellObj.value)
#     print('---Koniec wiersza---')

list(sheet.columns)  # List sheet column
#
for cellObj in list(sheet.columns)[0]:
    print(cellObj.value)